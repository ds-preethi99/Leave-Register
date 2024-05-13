import requests
import os
from datetime import datetime
import shutil
from openpyxl import load_workbook
import json
import calendar
import send_message
import pandas as pd
from configparser import ConfigParser

# Read the config file
config = ConfigParser()
config.read('config.ini')
# Read the Discord section from the config file
discord_section = config['Discord']
# Get values from the config file
endpoint = discord_section.get('ENDPOINT', '')
token = discord_section.get('TOKEN', '')
webhook_url = discord_section.get('WEBHOOK_URL', '')


def get_column_letter_from_coordinate(coordinate):
    return ''.join(filter(str.isalpha, coordinate))


# Send whatsapp message
def send_whatsapp_message_to_individual(logger, message, emp_manager_name):
    # API endpoint and token
    logger.info('send whatsapp message to individual method executed')
    message = ', '.join(message)
    logger.info(f'{message}')
    current_month = datetime.today().month  # get current month using the today().month function
    monthname = calendar.month_name[current_month]  # get current month name
    file_path = f'C:/Users/dsautomationslive/Employee Contact/{monthname}/Employee Contact.xlsx'
    employee_contact_book = load_workbook(file_path)
    employee_contact_sheet = employee_contact_book.active
    logger.info(f'get excel sheet {file_path}')
    manager_number = ''
    column_letter = get_column_letter_from_coordinate(employee_contact_sheet.cell(row=1, column=2).coordinate)
    for cell in employee_contact_sheet[column_letter]:
        if employee_contact_sheet.cell(row=cell.row, column=2).value == emp_manager_name:
            logger.info('get manager number')
            manager_number = employee_contact_sheet.cell(row=cell.row, column=3).value
            break
    logger.info(f'manager number {manager_number} \n {message}')
    # Message data
    data = {
        "messaging_product": "whatsapp",
        "recipient_type": "individual",
        "to": f"+91 {manager_number}",  # replace f"+91 {manager_number}"
        "type": "template",
        "template": {
            "name": "leave_register",
            "language": {
                "code": "en_US"
            },
            "components": [
                {
                    "type": "body",
                    "parameters": [
                        {"type": "text", "text": f"{emp_manager_name}"},
                        {"type": "text", "text": f"{message}"}
                    ]
                }
            ]
        }
    }

    headers = {
        'Authorization': f"Bearer {token}",
        'Content-Type': 'application/json'
    }

    response = requests.post(endpoint, headers=headers, data=json.dumps(data))

    if response.status_code == 200:
        logger.info(f'whatsapp message sent successfully - {message}')
    else:
        logger.error(f'whatsapp message didn\'t send for {manager_number}, {message} \n {response.text}')
# Send a discord message to the user


def send_discord_message_to_user(logger, lead_webhook_url):
    try:
        # get the file path URL
        book = load_workbook('Leave Register/Leave Register.xlsx')
        sheet = book.active
        working_hours_df = pd.read_excel('Employee Working Hours/Employee Hour Report.xlsx')
        # To store the Leave applied employee details
        emp_id = ''
        emp_id_list = []
        emp_name_list = []
        date_list = []
        emp_manager_list = []
        # Iterate the Applied Satus column
        for cell in sheet[sheet.cell(row=1, column=18).column_letter]:
            # check whether the Applied Status is "Applied"
            if cell.value == 'Applied':
                # Get Employee id who has sent the
                logger.info(f'Get the list of employee details who are applied for leave')
                emp_id = sheet.cell(row=cell.row, column=1).value
                emp_id_list.append(emp_id)
                # Get Employee Name who has sent the request
                emp_name = sheet.cell(row=cell.row, column=2).value
                emp_name_list.append(emp_name)
                # Get Employee Manager Name who has sent the request
                emp_manager = sheet.cell(row=cell.row, column=5).value
                emp_manager_list.append(emp_manager)
                # Get the Date
                date = sheet.cell(row=cell.row, column=10).value
                date_list.append(date)
                logger.info(f'{emp_id} | {emp_name} | {emp_manager} | {date}')
        if len(emp_id) == 0:
            message = f'Good Morning Team, there are no leave records for approval today.'
            send_message.send_message_in_discord_channel(lead_webhook_url, message, logger)
            send_message.send_message_in_discord_channel(webhook_url, message, logger)
        try:
            # Organizing the data
            organized_data = {}
            for manager, employee, date in zip(emp_manager_list, emp_name_list, date_list):
                if manager not in organized_data:
                    organized_data[manager] = {}
                if employee not in organized_data[manager]:
                    organized_data[manager][employee] = set()
                organized_data[manager][employee].add(date)
            # Formatting the message
            messages = []
            manager_name = []
            for manager, employees in organized_data.items():
                employee_messages = []
                for employee, dates in employees.items():
                    dates_list = sorted(list(dates))
                    # Check if the employee has multiple dates, and format accordingly
                    if len(dates_list) > 1:
                        last_date = dates_list.pop()
                        dates_str = ', '.join(dates_list) + ' and ' + last_date
                    else:
                        dates_str = dates_list[0]
                    employee_messages.append(f"{employee} has applied for leave on {dates_str}")
                manager_name.append(manager)
                logger.info(f"{manager}, '--', {employee_messages}")
                send_whatsapp_message_to_individual(logger, employee_messages, manager)
                discord_id = working_hours_df.loc[working_hours_df['Employee Name'] == manager, "Discord ID"].iloc[0]
                messages.append(f"Hi <@{discord_id}>, \n{', '.join(employee_messages)}. Please approve the leave.")
            emp_leave_list = zip(messages, manager_name)
            # Print the messages
            for message, manager in emp_leave_list:
                send_message.send_message_in_discord_channel(webhook_url, message, logger)
                send_message.send_message_in_discord_channel(lead_webhook_url, message, logger)
        except Exception as e:
            logger.info(f'Whatsapp message sent but getting split error')
            if "'list' object has no attribute 'split'" in str(e):
                send_message.send_message_in_discord_channel(webhook_url,
                                                             "Trying to split a list, "
                                                             "which is not supported.", logger)
            else:
                logger.critical(f'Error occur when sending whatsapp message')
                send_message.send_message_in_discord_channel(webhook_url,
                                                             f'Error occur in the send '
                                                             f'_discord_message_to_user method {e}', logger)
    except Exception as e:
        logger.critical(f'Error occur in the send _discord_message_to_user method {e}')
        send_message.send_message_in_discord_channel(webhook_url,
                                                     f'Error occur in the send _discord_message_to_user'
                                                     f' method {e}', logger)


def remove_current_file(file):
    if os.path.exists(file):
        os.remove(file)


# Create a Sub Folder Named Leave Register for today's date and moving the Leave Register file from
# one location to another
def create_sub_folder(logger, lead_webhook_url):
    try:
        file_path = 'C:/Users/dsautomationslive/Downloads/Leave Register.xlsx'
        if os.path.exists(file_path):
            logger.info(f'Leave Register file exist in the download folder')
            if os.path.exists(rf'Leave Register/Leave Register.xlsx'):
                remove_current_file('Leave Register/Leave Register.xlsx')
            shutil.move(file_path, rf'Leave Register/Leave Register.xlsx')
            logger.info(f'File moved from download path to Leave Register folder path ')
            # call the send_discord_message_to_user() function to perform the discord webhook
            send_discord_message_to_user(logger, lead_webhook_url)
        else:
            if os.path.exists(rf'Leave Register/Leave Register.xlsx'):
                send_discord_message_to_user(logger, lead_webhook_url)
            else:
                send_message.send_message_in_discord_channel(webhook_url,
                                                             f'Leave Register file does not exist so'
                                                             f' execution stopped', logger)
    except Exception as e:
        logger.critical(f'Error occur in the getValuesFromExcel.create_sub_folder {e}')
        send_message.send_message_in_discord_channel(webhook_url,
                                                     f'Error occurs in the getValuesFromExcel.'
                                                     f'create_sub_folder method {e}', logger)
# end of the create_sub_folder() method


# Check whether the file exist or not
def check_folder_exist_or_not(logger, lead_webhook_url):
    try:
        if not (os.path.exists('Leave Register')):
            os.makedirs('Leave Register')
        create_sub_folder(logger, lead_webhook_url)
        logger.info(f'Leave Register folder exist')
    except Exception as e:
        logger.critical(f'Error occurs in the check_folder_exist_or_not method {e}')
        send_message.send_message_in_discord_channel(webhook_url,
                                                     f'Error occurs in the check_folder_exist_or_not'
                                                     f' method {e}', logger)
# end of the check_folder_Exist_or_not() method
