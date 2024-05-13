import getValuesFromExcel
import os
from datetime import datetime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from time import sleep
import sys
from configparser import ConfigParser
import send_message
import log_file


# Read the config file
config = ConfigParser()
config.read('config.ini')
# Read the Discord section from the config file
discord_section = config['Discord']
# Get values from the config file
hrapp_user_id = discord_section.get('USER_ID', '')
hrapp_password = discord_section.get('PASSWORD', '')
webhook_url = discord_section.get('WEBHOOK_URL', '')
lead_webhook_url = discord_section.get('LEAD_WEBHOOK_URL', '')
health_checkup_webhook_url = discord_section.get('WEBHOOK_HEALTH_CHECKUP_URL', '')


today = datetime.today()

logfile = f'log/LeaveRegister_export{today.date()}.log'
logger = log_file.setup_logger(logfile)


def download_leave_register():
    try:
        logger.info("exportFileUsingSelenium started execution")
        # Create a new driver instance and add the webdriver
        s = Service('chromedriver.exe')
        driver = webdriver.Chrome(service=s)
        # Open HRApp
        driver.get('https://digitalseo.hrapp.co/auth')
        file_path = f'C:/Users/dsautomationslive/Company Holiday List {today.year}'
        if not (os.path.exists(file_path)):
            logger.info(f'Company Holiday List {today.year} folder not exist')
            send_message.send_message_in_discord_channel(health_checkup_webhook_url,
                                                         f'Company Holiday List {today.year} folder '
                                                         f'not exist in leave register folder', logger)
            sys.exit()
        # Splitting the date to get day, month, and year
        holiday_book = load_workbook(f'{file_path}/Location(s) Holidays.xlsx')  # get the activity tracker Excel sheet
        holiday_sheet = holiday_book.active  # set the activity_tracker_book sheet as active
        for cell in holiday_sheet[holiday_sheet.cell(row=1, column=4).column_letter]:
            if cell.value == today.date():
                logger.info(f'Leave Register execution stop due to holiday')
                send_message.send_message_in_discord_channel(health_checkup_webhook_url,
                                                             f'Leave Register execution stop due to holiday', logger)
                sys.exit()
        today_week = today.isoweekday()
        if today_week == 7:
            logger.info(f'Leave Register execution stop due to is sunday')
            send_message.send_message_in_discord_channel(health_checkup_webhook_url,
                                                         f'Leave Register execution stop due to sunday', logger)
            driver.quit()
            sys.exit()
        # Maximizing the Chrome window
        driver.maximize_window()
        sleep(8)
        # click on the sign with email button
        driver.find_element(By.CSS_SELECTOR, '#signinWithEmailBtnText').click()
        # Identify and interact with the username field
        driver.find_element(By.CSS_SELECTOR, '#formSigninEmailId').send_keys(hrapp_user_id)
        sleep(1)
        # click on the Next button
        driver.find_element(By.CSS_SELECTOR, '#email-verification-button').click()
        sleep(3)
        # Identify and interact with the password field
        driver.find_element(By.ID, 'formSigninPassword').send_keys(hrapp_password)
        sleep(2)
        # Identify and click the signin button
        driver.find_element(By.ID, 'email-password-submit-button').click()
        sleep(20)
        driver.get('https://digitalseo.hrapp.co/employees/leaves')
        sleep(10)
        # exportLeave button
        driver.find_element(By.ID, 'exportLeave').click()
        sleep(100)
        logger.info(f'Leave register file downloaded')
        # quit the program
        driver.quit()
        getValuesFromExcel.check_folder_exist_or_not(logger, lead_webhook_url)
    except Exception as e:
        logger.critical(f'Error occurs in the leave register {e}')
        send_message.send_message_in_discord_channel(health_checkup_webhook_url,
                                                     'Exception occurs in the Leave register script '
                                                     'kindly check the log', logger)
    finally:
        logger.info(f'Leave register executed successfully')
        send_message.send_message_in_discord_channel(health_checkup_webhook_url,
                                                     'Leave register executed successfully', logger)


def delete_leave_register_files(folder_path):
    for folder in folder_path:
        # List all files in the specified folder
        all_files = os.listdir(folder)
        # Filter files that contain "Leave Register" in their name
        leave_register_files = [file for file in all_files if "Leave Register" in file]
        # Delete each leave register file
        for file in leave_register_files:
            file_path = os.path.join(folder, file)
            os.remove(file_path)
            logger.info(f"Deleted: {file_path}")


# Example usage
download_folder_path = ['C:/Users/dsautomationslive/Downloads', 'Leave Register']
delete_leave_register_files(download_folder_path)
download_leave_register()
